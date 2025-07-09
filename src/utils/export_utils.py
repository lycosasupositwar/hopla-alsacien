import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

class ExportUtils:
    def __init__(self):
        pass
        
    def export_to_excel(self, results, file_path):
        """Export analysis results to Excel file"""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Export grain analysis
        if 'grain_analysis' in results:
            self.export_grain_analysis(wb, results['grain_analysis'])
            
        # Export phase analysis
        if 'phase_analysis' in results:
            self.export_phase_analysis(wb, results['phase_analysis'])
            
        # Export summary
        self.export_summary(wb, results)
        
        wb.save(file_path)
        
    def export_grain_analysis(self, wb, grain_data):
        """Export grain analysis to Excel sheet"""
        ws = wb.create_sheet("Grain Analysis")
        
        # Headers
        headers = ["ID", "Area (μm²)", "Equivalent Diameter (μm)", 
                  "Perimeter (μm)", "Shape Factor", "Orientation (°)"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            
        # Data
        grains = grain_data.get('grains', [])
        for row, grain in enumerate(grains, 2):
            ws.cell(row=row, column=1, value=grain.get('id', ''))
            ws.cell(row=row, column=2, value=round(grain.get('area', 0), 2))
            ws.cell(row=row, column=3, value=round(grain.get('equivalent_diameter', 0), 2))
            ws.cell(row=row, column=4, value=round(grain.get('perimeter', 0), 2))
            ws.cell(row=row, column=5, value=round(grain.get('shape_factor', 0), 3))
            ws.cell(row=row, column=6, value=round(grain.get('orientation', 0), 1))
            
        # ASTM summary
        summary_row = len(grains) + 3
        ws.cell(row=summary_row, column=1, value="ASTM Grain Size Number:").font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=grain_data.get('astm_number', 'N/A'))
        
        ws.cell(row=summary_row + 1, column=1, value="Mean Diameter:").font = Font(bold=True)
        ws.cell(row=summary_row + 1, column=2, value=f"{grain_data.get('mean_diameter', 0):.2f} μm")
        
        ws.cell(row=summary_row + 2, column=1, value="Total Grains:").font = Font(bold=True)
        ws.cell(row=summary_row + 2, column=2, value=len(grains))
        
    def export_phase_analysis(self, wb, phase_data):
        """Export phase analysis to Excel sheet"""
        ws = wb.create_sheet("Phase Analysis")
        
        # Headers
        headers = ["Phase", "Area (pixels)", "Percentage (%)", "Number of Regions"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            
        # Data
        phases = phase_data.get('phases', [])
        for row, phase in enumerate(phases, 2):
            ws.cell(row=row, column=1, value=phase.get('name', ''))
            ws.cell(row=row, column=2, value=phase.get('area', 0))
            ws.cell(row=row, column=3, value=round(phase.get('percentage', 0), 2))
            ws.cell(row=row, column=4, value=phase.get('num_regions', 0))
            
    def export_summary(self, wb, results):
        """Export analysis summary"""
        ws = wb.create_sheet("Summary")
        
        row = 1
        ws.cell(row=row, column=1, value="MetalloBox Analysis Summary").font = Font(bold=True, size=16)
        row += 2
        
        if 'grain_analysis' in results:
            grain_data = results['grain_analysis']
            ws.cell(row=row, column=1, value="Grain Analysis:").font = Font(bold=True)
            row += 1
            ws.cell(row=row, column=1, value=f"ASTM Grain Size Number: {grain_data.get('astm_number', 'N/A')}")
            row += 1
            ws.cell(row=row, column=1, value=f"Mean Diameter: {grain_data.get('mean_diameter', 0):.2f} μm")
            row += 1
            ws.cell(row=row, column=1, value=f"Total Grains: {len(grain_data.get('grains', []))}")
            row += 2
            
        if 'phase_analysis' in results:
            phase_data = results['phase_analysis']
            ws.cell(row=row, column=1, value="Phase Analysis:").font = Font(bold=True)
            row += 1
            ws.cell(row=row, column=1, value=f"Total Phases: {len(phase_data.get('phases', []))}")
            row += 1
            
            for phase in phase_data.get('phases', []):
                ws.cell(row=row, column=1, value=f"{phase.get('name', '')}: {phase.get('percentage', 0):.1f}%")
                row += 1
